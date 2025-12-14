from __future__ import annotations

from pathlib import Path
import tempfile
import unittest
import re

from openrouter_free_model_scouter.excel_timeline_repository import (
    ExcelTimelineRepository,
)
from openpyxl import load_workbook


class TestExcelBootstrapFromCsv(unittest.TestCase):
    def test_bootstrap_from_csv_if_needed_creates_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            base_dir = Path(tmp_dir)
            csv_path = base_dir / "history.csv"
            xlsx_path = base_dir / "history.xlsx"

            csv_path.write_text(
                "timestamp_iso,run_id,model_id,ok,http_status,latency_ms,attempts,error_category,error_message,response_preview\n"
                "2025-01-01T00:00:00+00:00,run1,a:free,true,200,10,1,,,OK\n"
                "2025-01-01T00:00:00+00:00,run1,b:free,false,429,20,2,rate_limited,rate limit,\n",
                encoding="utf-8",
            )

            repository = ExcelTimelineRepository()
            repository.bootstrap_from_csv_if_needed(
                xlsx_path=xlsx_path, csv_path=csv_path
            )

            self.assertTrue(xlsx_path.exists())

            workbook = load_workbook(xlsx_path)
            worksheet = workbook.active
            header_value = worksheet.cell(row=1, column=2).value
            self.assertIsNotNone(header_value)
            self.assertIsInstance(header_value, str)
            self.assertRegex(
                header_value, re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$")
            )


if __name__ == "__main__":
    unittest.main()
