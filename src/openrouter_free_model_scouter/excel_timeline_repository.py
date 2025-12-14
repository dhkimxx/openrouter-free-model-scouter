from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel, to_excel

from .domain_models import HealthcheckResult


@dataclass(frozen=True)
class TimelineCell:
    value: str
    fill: Optional[PatternFill]


class ExcelTimelineRepository:
    def append_run(
        self,
        xlsx_path: Path,
        *,
        run_datetime: datetime,
        results: Iterable[HealthcheckResult],
    ) -> None:
        header_text = self._format_run_datetime_for_header(run_datetime)

        xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        workbook, worksheet = self._open_or_create(xlsx_path)
        self._normalize_existing_cells_for_display(workbook, worksheet)

        header_row_index = 1
        model_id_column_index = 1
        run_column_index = worksheet.max_column + 1
        if (
            worksheet.max_row == 1
            and worksheet.max_column == 1
            and worksheet["A1"].value is None
        ):
            worksheet.cell(
                row=header_row_index, column=model_id_column_index, value="model_id"
            )
            run_column_index = 2

        header_cell = worksheet.cell(
            row=header_row_index, column=run_column_index, value=header_text
        )
        header_cell.number_format = "@"
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center")

        worksheet.freeze_panes = "B2"

        model_row_map = self._build_model_row_map(worksheet)

        for result in results:
            row_index = model_row_map.get(result.model_id)
            if row_index is None:
                row_index = worksheet.max_row + 1
                worksheet.cell(
                    row=row_index, column=model_id_column_index, value=result.model_id
                ).font = Font(bold=True)
                model_row_map[result.model_id] = row_index

            timeline_cell = self._format_timeline_cell(result)
            cell = worksheet.cell(
                row=row_index, column=run_column_index, value=timeline_cell.value
            )
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "@"
            if timeline_cell.fill is not None:
                cell.fill = timeline_cell.fill

        self._apply_basic_layout(worksheet)
        workbook.save(xlsx_path)

    def bootstrap_from_csv_if_needed(
        self,
        *,
        xlsx_path: Path,
        csv_path: Path,
    ) -> None:
        if xlsx_path.exists():
            return
        if not csv_path.exists():
            return

        runs = self._read_csv_runs(csv_path)
        if not runs:
            return

        for run_datetime, results in runs:
            self.append_run(xlsx_path, run_datetime=run_datetime, results=results)

    def _open_or_create(self, xlsx_path: Path):
        if xlsx_path.exists():
            workbook = load_workbook(xlsx_path)
            worksheet = workbook.active
            return workbook, worksheet

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "timeline"
        return workbook, worksheet

    def _build_model_row_map(self, worksheet) -> Dict[str, int]:
        model_row_map: Dict[str, int] = {}
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=1).value
            if isinstance(value, str) and value:
                model_row_map[value] = row_index
        return model_row_map

    def _format_timeline_cell(self, result: HealthcheckResult) -> TimelineCell:
        ok_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        rate_limit_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        fail_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        if result.ok:
            latency_text = (
                "" if result.latency_ms is None else f" ({result.latency_ms}ms)"
            )
            return TimelineCell(value=f"OK{latency_text}", fill=ok_fill)

        if result.http_status == 429 or result.error_category == "rate_limited":
            return TimelineCell(value="429", fill=rate_limit_fill)

        if result.http_status is not None:
            return TimelineCell(value=f"HTTP {result.http_status}", fill=fail_fill)

        if result.error_category:
            return TimelineCell(value=result.error_category, fill=fail_fill)

        return TimelineCell(value="FAIL", fill=fail_fill)

    def _apply_basic_layout(self, worksheet) -> None:
        header_fill = PatternFill(
            start_color="1F2937", end_color="1F2937", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        model_fill = PatternFill(
            start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"
        )

        worksheet.column_dimensions["A"].width = 45

        header_row_index = 1
        for col_index in range(2, worksheet.max_column + 1):
            column_letter = worksheet.cell(
                row=header_row_index, column=col_index
            ).column_letter
            worksheet.column_dimensions[column_letter].width = 18
            worksheet.cell(row=header_row_index, column=col_index).number_format = "@"

        for col_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_index)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=1)
            if cell.value:
                cell.fill = model_fill
                cell.alignment = Alignment(horizontal="left")

    def _read_csv_runs(
        self, csv_path: Path
    ) -> List[Tuple[datetime, List[HealthcheckResult]]]:
        rows: List[Dict[str, str]] = []
        with csv_path.open("r", encoding="utf-8", newline="") as file:
            reader = csv.DictReader(file)
            for row in reader:
                rows.append({k: (v or "") for k, v in row.items()})

        grouped: Dict[str, List[Dict[str, str]]] = {}
        for row in rows:
            run_id = row.get("run_id", "")
            if not run_id:
                continue
            grouped.setdefault(run_id, []).append(row)

        runs: List[Tuple[datetime, List[HealthcheckResult]]] = []
        for run_id, run_rows in grouped.items():
            datetimes: List[datetime] = []
            results: List[HealthcheckResult] = []

            for row in run_rows:
                timestamp_iso = row.get("timestamp_iso", "")
                parsed_datetime = self._try_parse_iso_datetime(timestamp_iso)
                if parsed_datetime is not None:
                    datetimes.append(parsed_datetime)

                model_id = row.get("model_id", "")
                ok = row.get("ok", "").lower() == "true"

                http_status = self._try_parse_int(row.get("http_status", ""))
                latency_ms = self._try_parse_int(row.get("latency_ms", ""))
                attempts = self._try_parse_int(row.get("attempts", "")) or 0

                error_category = row.get("error_category", "") or None
                error_message = row.get("error_message", "") or None
                response_preview = row.get("response_preview", "") or None

                if not model_id:
                    continue

                results.append(
                    HealthcheckResult(
                        run_id=run_id,
                        timestamp_iso=timestamp_iso,
                        model_id=model_id,
                        ok=ok,
                        http_status=http_status,
                        latency_ms=latency_ms,
                        attempts=attempts,
                        error_category=error_category,
                        error_message=error_message,
                        response_preview=response_preview,
                    )
                )

            if not results:
                continue

            run_datetime = min(datetimes) if datetimes else datetime.now()
            run_datetime = self._to_excel_header_datetime(run_datetime)
            runs.append((run_datetime, results))

        runs.sort(key=lambda item: item[0])
        return runs

    def _try_parse_iso_datetime(self, value: str) -> Optional[datetime]:
        if not value:
            return None
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None

    def _to_excel_header_datetime(self, value: datetime) -> datetime:
        if value.tzinfo is None:
            return value
        return value.astimezone().replace(tzinfo=None)

    def _format_run_datetime_for_header(self, value: datetime) -> str:
        normalized = self._to_excel_header_datetime(value)
        return normalized.strftime("%Y-%m-%d %H:%M:%S")

    def _normalize_existing_cells_for_display(self, workbook, worksheet) -> None:
        header_row_index = 1
        epoch = workbook.epoch
        for col_index in range(2, worksheet.max_column + 1):
            cell = worksheet.cell(row=header_row_index, column=col_index)
            if isinstance(cell.value, datetime):
                cell.value = self._format_run_datetime_for_header(cell.value)
            elif isinstance(cell.value, (int, float)):
                try:
                    cell.value = self._format_run_datetime_for_header(
                        from_excel(cell.value, epoch=epoch)
                    )
                except (TypeError, ValueError):
                    cell.value = str(cell.value)
            elif cell.value is not None and not isinstance(cell.value, str):
                cell.value = str(cell.value)
            cell.number_format = "@"

        for row_index in range(2, worksheet.max_row + 1):
            for col_index in range(2, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_index, column=col_index)
                if isinstance(cell.value, datetime):
                    try:
                        serial = to_excel(cell.value, epoch=epoch)
                        cell.value = str(int(serial))
                    except (TypeError, ValueError):
                        cell.value = str(cell.value)
                elif isinstance(cell.value, (int, float)) and not isinstance(
                    cell.value, bool
                ):
                    if isinstance(cell.value, float) and cell.value.is_integer():
                        cell.value = str(int(cell.value))
                    else:
                        cell.value = str(cell.value)
                elif cell.value is not None and not isinstance(cell.value, str):
                    cell.value = str(cell.value)
                cell.number_format = "@"

    def _try_parse_int(self, value: str) -> Optional[int]:
        if not value:
            return None
        try:
            return int(value)
        except ValueError:
            return None
