from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import load_workbook

from .constants import STATUS_MISSING, STATUS_OK, STATUS_RATE_LIMITED, STATUS_FAIL


def normalize_run_label(value: object, fallback_index: int) -> str:
    if value is None:
        return f"run-{fallback_index}"

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    text = str(value).strip()
    if not text:
        return f"run-{fallback_index}"

    # Keep timeline labels compact and markdown-safe.
    return text.replace("|", "/")


def parse_status(value: object) -> str:
    if value is None:
        return STATUS_MISSING

    text = str(value).strip()
    if not text:
        return STATUS_MISSING

    upper_text = text.upper()
    if upper_text.startswith(STATUS_OK):
        return STATUS_OK
    if text == STATUS_RATE_LIMITED:
        return STATUS_RATE_LIMITED

    return STATUS_FAIL


def read_timeline(xlsx_path: Path) -> Tuple[List[str], Dict[str, List[str]]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active

    if worksheet.max_column < 2:
        raise ValueError("timeline workbook does not contain any run columns")

    run_labels: List[str] = []
    for col_index in range(2, worksheet.max_column + 1):
        header_value = worksheet.cell(row=1, column=col_index).value
        run_labels.append(normalize_run_label(header_value, col_index - 1))

    model_statuses: Dict[str, List[str]] = {}
    for row_index in range(2, worksheet.max_row + 1):
        model_value = worksheet.cell(row=row_index, column=1).value
        if model_value is None:
            continue

        model_id = str(model_value).strip()
        if not model_id:
            continue

        statuses = [
            parse_status(worksheet.cell(row=row_index, column=col_index).value)
            for col_index in range(2, worksheet.max_column + 1)
        ]
        model_statuses[model_id] = statuses

    if not model_statuses:
        raise ValueError("timeline workbook does not contain model rows")

    return run_labels, model_statuses


def slice_lookback(
    run_labels: Sequence[str],
    model_statuses: Dict[str, List[str]],
    lookback_runs: int,
) -> Tuple[List[str], Dict[str, List[str]]]:
    if lookback_runs <= 0:
        raise ValueError("lookback-runs must be >= 1")

    run_count = len(run_labels)
    if lookback_runs >= run_count:
        return list(run_labels), {k: list(v) for k, v in model_statuses.items()}

    start_index = run_count - lookback_runs
    sliced_labels = list(run_labels[start_index:])
    sliced_statuses = {k: list(v[start_index:]) for k, v in model_statuses.items()}
    return sliced_labels, sliced_statuses
